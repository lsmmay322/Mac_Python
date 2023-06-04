import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl import load_workbook

def extract_data(ex_data, new_data, data_num):
	row_index = 2
	col_index = data_num*4-3
	row_new_index=2
	col_new_index=2*x-1
	while(1):
		data_ocv = ex_data.cell(row=row_index, column=col_index).value
		if data_ocv == None:
			break
		data_DCIR = ex_data.cell(row=row_index, column=col_index+2).value
		new_data.cell(row=row_new_index, column=col_new_index, value=data_ocv)
		new_data.cell(row=row_new_index, column=col_new_index+1, value=data_DCIR)
		row_new_index = row_new_index + 1
	return
	
# Gui로 엑셀 파일 불러오기
root = tk.Tk()

root.title("엑셀 파일 열기")

root.geometry("1000x400+100+100")

filename =tk.filedialog.askopenfilename(initialdir = "~/Coding",
    title = "open file", filetypes = (("*.xlsx", "*xlsx"), ("*.xls", "*xls"), ("*.csv", "*csv")))

if filename == '':
    messagebox.showwarning("경고", "파일을 추가하세요")
    
# DCIR 엑셀 파일 불러오기
load_wb = load_workbook(filename, data_only=True)
load_ws = load_wb['Sheet1']

# OCV와 DCIR 데이터 옮겨 담을 엑셀 파일 생성
write_wb = Workbook()
write_ws = write_wb.create_sheet('DCIR')
write_ws = write_wb.active

x = 1
while(1):
	data = load_ws.cell(row=1, column=4*x-3).value
	if data == None:
		break
	write_ws.cell(row=1, column=2*x-1, value=data) # 데이터 이름 옮겨담기
	extract_data(load_ws,write_ws, x) # 데이터 자체 옮겨담기
	x = x + 1

write_wb.save("/Users/weed/Coding/Python/SE/DCIR.xlsx")

    
